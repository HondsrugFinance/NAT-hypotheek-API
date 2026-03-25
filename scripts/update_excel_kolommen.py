"""Update Excel: hernoem Opmerkingen, voeg Auto/Handmatig toe, herschik kolommen."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from copy import copy

wb = openpyxl.load_workbook("docs/document-extractie-mapping.xlsx")

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2E5644", end_color="2E5644", fill_type="solid")

tabs_to_update = [
    "Persoonsgegevens",
    "Inkomen",
    "Onderpand & Financiering",
    "Verplichtingen & Hypotheken",
    "Voorzieningen & Vermogen",
]

for sheet_name in tabs_to_update:
    ws = wb[sheet_name]
    max_col = ws.max_column
    max_row = ws.max_row

    # Vind kolom-indices op basis van header-namen
    headers = {}
    for c in range(1, max_col + 1):
        val = str(ws.cell(row=1, column=c).value or "")
        headers[val] = c

    # Bepaal huidige posities
    opmerkingen_col = headers.get("Opmerkingen")
    validatie_col = headers.get("Validatieregel")

    # Voeg "Automatisch/Handmatig" kolom toe na de bron-kolommen, voor Validatieregel
    # Gewenste volgorde: Veld | Type | Bron 1-4 | Auto/Handmatig | Validatieregel | Notities adviseur

    # Stap 1: Lees alle data uit
    all_data = []
    for r in range(1, max_row + 1):
        row_data = {}
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            row_data[c] = {
                "value": cell.value,
                "font": copy(cell.font) if cell.font else None,
                "fill": copy(cell.fill) if cell.fill else None,
                "alignment": copy(cell.alignment) if cell.alignment else None,
            }
        all_data.append(row_data)

    # Stap 2: Bepaal gewenste kolomvolgorde
    # Zoek alle kolom-headers
    header_row = all_data[0]
    col_names = {c: str(header_row[c]["value"] or "") for c in header_row}

    # Gewenste volgorde
    desired_order = []
    # Eerst: Veld
    for c, name in col_names.items():
        if name == "Veld":
            desired_order.append(c)
            break
    # Type
    for c, name in col_names.items():
        if name.startswith("Type"):
            desired_order.append(c)
            break
    # Bron document 1-4
    for i in range(1, 5):
        for c, name in col_names.items():
            if name == f"Bron document {i}" or (i == 1 and name == "Bron document"):
                desired_order.append(c)
                break
    # De rest die nog niet is toegevoegd (behalve Opmerkingen en Validatieregel)
    skip_names = {"Veld", "Type / Opties", "Opmerkingen", "Validatieregel"}
    for c, name in col_names.items():
        if c not in desired_order and name not in skip_names and name.startswith("Bron"):
            desired_order.append(c)
    # Validatieregel
    if validatie_col:
        desired_order.append(validatie_col)
    # Opmerkingen (wordt Notities adviseur) - altijd laatste
    if opmerkingen_col:
        desired_order.append(opmerkingen_col)

    # Stap 3: Wis het werkblad en schrijf opnieuw in gewenste volgorde
    # Plus voeg Auto/Handmatig toe

    # Maak nieuw werkblad-data
    new_headers = []
    source_cols = []
    for c in desired_order:
        name = col_names.get(c, "")
        if name == "Opmerkingen":
            name = "Notities adviseur"
        new_headers.append(name)
        source_cols.append(c)

    # Voeg Auto/Handmatig in na de laatste Bron kolom, voor Validatieregel
    insert_pos = len(new_headers)  # default: einde
    for i, name in enumerate(new_headers):
        if name == "Validatieregel":
            insert_pos = i
            break
    new_headers.insert(insert_pos, "Auto / Handmatig")
    source_cols.insert(insert_pos, None)  # None = nieuwe kolom

    # Wis alle cellen
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 5):
            ws.cell(row=r, column=c).value = None
            ws.cell(row=r, column=c).font = Font()
            ws.cell(row=r, column=c).fill = PatternFill()
            ws.cell(row=r, column=c).alignment = Alignment()

    # Schrijf headers
    for i, name in enumerate(new_headers, 1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    # Schrijf data
    for r_idx, row_data in enumerate(all_data[1:], 2):  # skip header
        for new_c, src_c in enumerate(source_cols, 1):
            if src_c is None:
                # Nieuwe kolom Auto/Handmatig - laat leeg voor nu
                continue
            if src_c in row_data:
                d = row_data[src_c]
                cell = ws.cell(row=r_idx, column=new_c, value=d["value"])
                if d["font"]:
                    cell.font = d["font"]
                if d["fill"]:
                    cell.fill = d["fill"]
                if d["alignment"]:
                    cell.alignment = d["alignment"]

    # Kolombreedtes
    for i, name in enumerate(new_headers, 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        if name == "Veld":
            ws.column_dimensions[col_letter].width = 30
        elif name.startswith("Type"):
            ws.column_dimensions[col_letter].width = 40
        elif name.startswith("Bron"):
            ws.column_dimensions[col_letter].width = 25
        elif name == "Auto / Handmatig":
            ws.column_dimensions[col_letter].width = 20
        elif name == "Validatieregel":
            ws.column_dimensions[col_letter].width = 50
        elif name == "Notities adviseur":
            ws.column_dimensions[col_letter].width = 40

# Vul Auto/Handmatig in voor Persoonsgegevens
ws = wb["Persoonsgegevens"]
# Zoek de kolom
auto_col = None
for c in range(1, ws.max_column + 1):
    if str(ws.cell(row=1, column=c).value or "") == "Auto / Handmatig":
        auto_col = c
        break

if auto_col:
    auto_mapping = {
        "geslacht": "Auto",
        "voorletters": "Auto",
        "voornamen": "Auto",
        "roepnaam": "Handmatig",
        "tussenvoegsel": "Auto",
        "achternaam": "Auto",
        "geboortedatum": "Auto",
        "geboorteplaats": "Auto",
        "geboorteland": "Auto",
        "nationaliteit": "Auto",
        "eerderGehuwd": "Auto",
        "datumEchtscheiding": "Auto",
        "weduweWeduwnaar": "Handmatig",
        "postcode": "Handmatig",
        "huisnummer": "Handmatig",
        "toevoeging": "Handmatig",
        "straat": "Auto (Postcode API)",
        "woonplaats": "Auto (Postcode API)",
        "land": "Auto (default)",
        "email": "Handmatig",
        "telefoonnummer": "Handmatig",
        "legitimatiesoort": "Auto",
        "legitimatienummer": "Auto",
        "afgiftedatum": "Auto",
        "geldigTot": "Auto",
        "afgifteplaats": "Auto",
        "afgifteland": "Auto",
        "heeftPartner": "Handmatig",
        "burgerlijkeStaat": "Auto / bevestigen",
        "samenlevingsvorm": "Auto / bevestigen",
        "kinderen[].geboortedatum": "Handmatig",
        "kinderen[].roepnaam": "Handmatig",
    }
    for row in ws.iter_rows(min_row=2, max_col=1):
        veld = str(row[0].value or "")
        if veld in auto_mapping:
            ws.cell(row=row[0].row, column=auto_col, value=auto_mapping[veld])

# Vul Auto/Handmatig in voor Inkomen
ws2 = wb["Inkomen"]
auto_col2 = None
for c in range(1, ws2.max_column + 1):
    if str(ws2.cell(row=1, column=c).value or "") == "Auto / Handmatig":
        auto_col2 = c
        break

if auto_col2:
    auto_mapping2 = {
        "naamWerkgever": "Auto",
        "postcodeWerkgever": "Auto",
        "adresWerkgever": "Auto",
        "vestigingsplaats": "Auto",
        "kvkNummer": "Auto",
        "soortBerekening": "Handmatig",
        "beroepstype": "Auto",
        "functie": "Auto",
        "soortDienstverband": "Auto",
        "gemiddeldUrenPerWeek": "Auto",
        "inDienstSinds": "Auto",
        "proeftijd": "Auto",
        "proeftijdVerstreken": "Auto",
        "einddatumContract": "Auto",
        "loonbeslag": "Auto",
        "onderhandseLening": "Auto",
        "directeurAandeelhouder": "Auto",
        "dienstbetrekkingBijFamilie": "Auto",
        "brutoSalaris": "Auto",
        "periode": "Auto",
        "vakantiegeldPercentage": "Auto",
        "vakantiegeldBedrag": "Auto",
        "eindejaarsuitkering": "Auto",
        "onregelmatigheidstoeslag": "Auto",
        "overwerk": "Auto",
        "provisie": "Auto",
        "structureelFlexibelBudget": "Auto",
        "vebAfgelopen12Maanden": "Auto",
        "dertiendeMaand": "Auto",
        "variabelBrutoJaarinkomen": "Auto",
        "vastToeslagOpHetInkomen": "Auto",
        "gemiddeldJaarToetsinkomen": "Auto (IBL-tool)",
        "maandelijksePensioenbijdrage": "Auto",
        "aantalWerkgevers": "Auto",
        "jaar1": "Auto",
        "jaar2": "Auto",
        "jaar3": "Auto",
        "soort": "Handmatig",
        "rekenmethode": "Handmatig",
        "nettoWinstJaar1": "Auto",
        "nettoWinstJaar2": "Auto",
        "nettoWinstJaar3": "Auto",
        "bedrijfsnaam": "Auto",
        "type": "Auto",
        "rechtsvorm": "Auto",
        "startdatumOnderneming": "Auto",
        "ouderdomspensioen.bedrag": "Auto",
        "ouderdomspensioen.ingangsdatum": "Auto",
        "ouderdomspensioen.standPer": "Auto",
        "partnerpensioen.verzekerdVoor": "Auto",
        "wezenpensioen.verzekerd": "Auto",
        "soortUitkering": "Auto",
        "startdatum": "Auto",
        "einddatum": "Auto",
        "soortAnderInkomen": "Handmatig",
    }
    for row in ws2.iter_rows(min_row=2, max_col=1):
        veld = str(row[0].value or "")
        if veld in auto_mapping2:
            ws2.cell(row=row[0].row, column=auto_col2, value=auto_mapping2[veld])

# Vul Auto/Handmatig in voor Onderpand
ws3 = wb["Onderpand & Financiering"]
auto_col3 = None
for c in range(1, ws3.max_column + 1):
    if str(ws3.cell(row=1, column=c).value or "") == "Auto / Handmatig":
        auto_col3 = c
        break

if auto_col3:
    auto_mapping3 = {
        "postcode": "Auto",
        "huisnummer": "Auto",
        "straat": "Auto",
        "woonplaats": "Auto",
        "typeWoning": "Auto",
        "soortOnderpand": "Auto",
        "marktwaarde": "Auto",
        "marktwaardeNaVerbouwing": "Auto",
        "wozWaarde": "Auto",
        "energielabel": "Auto (EP-Online)",
        "erfpacht": "Auto",
        "jaarlijkseErfpacht": "Auto",
        "aankoopsomWoning": "Auto",
        "overdrachtsbelastingPercentage": "Handmatig",
        "verbouwing": "Auto",
        "ebv": "Handmatig",
        "taxatiekosten": "Handmatig",
        "eigenGeld": "Auto / bevestigen",
        "koopprijs": "Auto",
        "leveringsdatum": "Auto",
        "datumBankgarantie": "Auto",
        "ontbindendeVoorwaardenDatum": "Auto",
    }
    for row in ws3.iter_rows(min_row=2, max_col=1):
        veld = str(row[0].value or "")
        if veld in auto_mapping3:
            ws3.cell(row=row[0].row, column=auto_col3, value=auto_mapping3[veld])

# Vul Auto/Handmatig in voor Verplichtingen
ws4 = wb["Verplichtingen & Hypotheken"]
auto_col4 = None
for c in range(1, ws4.max_column + 1):
    if str(ws4.cell(row=1, column=c).value or "") == "Auto / Handmatig":
        auto_col4 = c
        break

if auto_col4:
    auto_mapping4 = {
        "type": "Auto",
        "kredietnummer": "Auto",
        "maatschappij": "Auto",
        "kredietbedrag": "Auto",
        "maandbedrag": "Auto",
        "saldo": "Auto",
        "nogAfTeLossen": "Auto",
        "rentepercentage": "Auto",
        "ingangsdatum": "Auto",
        "einddatum": "Auto",
        "status": "Handmatig",
        "geldverstrekker": "Auto",
        "inschrijving": "Auto",
        "rangorde": "Auto",
        "nhg": "Auto",
        "aflosvorm": "Auto",
        "bedrag": "Auto",
        "rentePercentage": "Auto",
        "looptijd": "Auto",
        "renteVastPeriode": "Auto",
        "einddatumRvp": "Auto",
        "restschuld": "Auto",
    }
    for row in ws4.iter_rows(min_row=2, max_col=1):
        veld = str(row[0].value or "")
        if veld in auto_mapping4:
            ws4.cell(row=row[0].row, column=auto_col4, value=auto_mapping4[veld])

# Vul Auto/Handmatig in voor Voorzieningen
ws5 = wb["Voorzieningen & Vermogen"]
auto_col5 = None
for c in range(1, ws5.max_column + 1):
    if str(ws5.cell(row=1, column=c).value or "") == "Auto / Handmatig":
        auto_col5 = c
        break

if auto_col5:
    auto_mapping5 = {
        "type": "Handmatig",
        "eigenaar": "Handmatig",
        "aanbieder": "Auto",
        "polisnummer": "Auto",
        "ingangsdatum": "Auto",
        "einddatum": "Auto",
        "soortDekking": "Auto",
        "premieBedrag": "Auto",
        "orvDekking": "Auto",
        "dekkingAO": "Auto",
        "dekkingWW": "Auto",
        "eersteJaar": "Auto",
        "tweedeJaar": "Auto",
        "ibanAanvrager": "Auto",
        "ibanPartner": "Auto",
        "spaargeld": "Auto",
        "beleggingen": "Auto",
        "schenking": "Handmatig",
    }
    for row in ws5.iter_rows(min_row=2, max_col=1):
        veld = str(row[0].value or "")
        if veld in auto_mapping5:
            ws5.cell(row=row[0].row, column=auto_col5, value=auto_mapping5[veld])

wb.save("docs/document-extractie-mapping.xlsx")
print("Klaar: Auto/Handmatig kolom + Notities adviseur (rechts) + herschikt")
