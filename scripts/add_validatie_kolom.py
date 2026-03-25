"""Voeg validatieregel kolom toe aan document-extractie-mapping.xlsx."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.load_workbook("docs/document-extractie-mapping.xlsx")

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2E5644", end_color="2E5644", fill_type="solid")

# Voeg kolom 'Validatieregel' toe aan tabblad 1-5
for sheet_name in [
    "Persoonsgegevens",
    "Inkomen",
    "Onderpand & Financiering",
    "Verplichtingen & Hypotheken",
    "Voorzieningen & Vermogen",
]:
    ws = wb[sheet_name]
    new_col = ws.max_column + 1
    cell = ws.cell(row=1, column=new_col, value="Validatieregel")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="left")
    col_letter = openpyxl.utils.get_column_letter(new_col)
    ws.column_dimensions[col_letter].width = 50


# --- Persoonsgegevens ---
ws = wb["Persoonsgegevens"]
col = ws.max_column
regels = {
    "geslacht": "Paspoort is leidend. UWV ter verificatie.",
    "voorletters": "Afleiden uit voornamen op paspoort (eerste letters + punten).",
    "voornamen": "Paspoort is leidend. Altijd voluit, nooit afkorten.",
    "roepnaam": "Handmatig invullen door adviseur of klant.",
    "tussenvoegsel": "Paspoort is leidend. Als alleen op UWV: meld aan adviseur.",
    "achternaam": "Paspoort is leidend. Bij afwijking UWV: waarschuwing.",
    "geboortedatum": "Moet identiek zijn op alle bronnen. Bij verschil: foutmelding.",
    "geboorteplaats": "Paspoort is leidend.",
    "geboorteland": "Paspoort is leidend.",
    "nationaliteit": "Paspoort is leidend.",
    "eerderGehuwd": "Alleen als echtscheidingsconvenant aanwezig.",
    "datumEchtscheiding": "Echtscheidingsconvenant is leidend.",
    "weduweWeduwnaar": "Handmatig. Niet automatisch afleidbaar.",
    "postcode": "Handmatig of uit loonstrook. Koopovereenkomst = nieuw adres (na verhuizing).",
    "huisnummer": "Handmatig of uit loonstrook. Koopovereenkomst = nieuw adres.",
    "toevoeging": "Zelfde als huisnummer.",
    "straat": "Afleiden via Postcode API op basis van postcode + huisnummer.",
    "woonplaats": "Afleiden via Postcode API op basis van postcode + huisnummer.",
    "email": "Bij voorkeur uit eerste contactmoment (e-mail afzender).",
    "telefoonnummer": "Handmatig of uit contacthistorie.",
    "legitimatiesoort": "Automatisch: paspoort of ID-kaart herkennen.",
    "legitimatienummer": "Paspoort/ID. Exacte match vereist (OCR check).",
    "afgiftedatum": "Paspoort/ID.",
    "geldigTot": "Paspoort/ID. Waarschuwing als <6 maanden geldig.",
    "afgifteplaats": "Paspoort/ID.",
    "afgifteland": "Paspoort/ID.",
    "burgerlijkeStaat": "Koopovereenkomst bevat vaak juridische status. Handmatig bevestigen.",
    "samenlevingsvorm": "Koopovereenkomst of handmatig.",
    "kinderen[].geboortedatum": "IB-aangifte of handmatig.",
    "kinderen[].roepnaam": "IB-aangifte of handmatig.",
}
for row in ws.iter_rows(min_row=2, max_col=1):
    veld = str(row[0].value or "")
    if veld in regels:
        c = ws.cell(row=row[0].row, column=col, value=regels[veld])
        c.alignment = Alignment(wrap_text=True, vertical="top")


# --- Inkomen ---
ws2 = wb["Inkomen"]
col2 = ws2.max_column
regels2 = {
    "naamWerkgever": "WGV is leidend. Loonstrook en UWV ter verificatie.",
    "soortBerekening": "Keuze adviseur. Systeem kan suggestie doen op basis van beschikbare docs.",
    "functie": "WGV is leidend.",
    "soortDienstverband": "WGV is leidend. Vast/tijdelijk/flexibel.",
    "gemiddeldUrenPerWeek": "WGV is leidend. Loonstrook ter controle.",
    "inDienstSinds": "WGV is leidend. Bij verschil met loonstrook: waarschuwing.",
    "einddatumContract": "WGV is leidend. Alleen bij tijdelijk dienstverband.",
    "brutoSalaris": "WGV is leidend. Controleer tegen loonstrook: verschil >5% = waarschuwing.",
    "vakantiegeldPercentage": "WGV is leidend. Default 8% als niet ingevuld.",
    "vakantiegeldBedrag": "WGV is leidend.",
    "eindejaarsuitkering": "WGV is leidend. Controleer of structureel.",
    "onregelmatigheidstoeslag": "WGV is leidend. Moet structureel zijn (>12 mnd).",
    "overwerk": "WGV is leidend. Gemiddelde afgelopen 3 jaar.",
    "provisie": "WGV is leidend. Gemiddelde afgelopen 3 jaar.",
    "dertiendeMaand": "WGV is leidend.",
    "gemiddeldJaarToetsinkomen": "Berekend door IBL-tool uit UWV. Niet handmatig wijzigen.",
    "maandelijksePensioenbijdrage": "Loonstrook is leidend. Input voor IBL-tool.",
    "nettoWinstJaar1": "Jaarrekening is leidend. IB-aangifte ter controle.",
    "nettoWinstJaar2": "Jaarrekening is leidend. IB-aangifte ter controle.",
    "nettoWinstJaar3": "Jaarrekening is leidend. IB-aangifte ter controle.",
    "ouderdomspensioen.bedrag": "Pensioenspecificatie/UPO is leidend.",
    "soortUitkering": "Toekenningsbesluit is leidend.",
    "jaarlijksBrutoInkomen": "Betaalspecificatie is leidend. Toekenningsbesluit ter controle.",
}
for row in ws2.iter_rows(min_row=2, max_col=1):
    veld = str(row[0].value or "")
    if veld in regels2:
        c = ws2.cell(row=row[0].row, column=col2, value=regels2[veld])
        c.alignment = Alignment(wrap_text=True, vertical="top")


# --- Onderpand ---
ws3 = wb["Onderpand & Financiering"]
col3 = ws3.max_column
regels3 = {
    "postcode": "Koopovereenkomst is leidend.",
    "huisnummer": "Koopovereenkomst is leidend.",
    "straat": "Koopovereenkomst is leidend. Postcode API ter verificatie.",
    "woonplaats": "Koopovereenkomst is leidend. Postcode API ter verificatie.",
    "marktwaarde": "Taxatierapport is leidend. Moet recenter zijn dan 6 maanden.",
    "marktwaardeNaVerbouwing": "Taxatierapport is leidend.",
    "wozWaarde": "WOZ-beschikking is leidend. EP-Online als alternatief.",
    "energielabel": "EP-Online API is leidend (altijd actueel). Document als fallback.",
    "aankoopsomWoning": "Koopovereenkomst is leidend. Moet exact overeenkomen.",
    "koopprijs": "Koopovereenkomst. Identiek aan aankoopsomWoning.",
    "leveringsdatum": "Koopovereenkomst is leidend.",
    "erfpacht": "Koopovereenkomst is leidend. Erfpachtakte voor details.",
}
for row in ws3.iter_rows(min_row=2, max_col=1):
    veld = str(row[0].value or "")
    if veld in regels3:
        c = ws3.cell(row=row[0].row, column=col3, value=regels3[veld])
        c.alignment = Alignment(wrap_text=True, vertical="top")


# --- Verplichtingen ---
ws4 = wb["Verplichtingen & Hypotheken"]
col4 = ws4.max_column
regels4 = {
    "type": "BKR is leidend voor type en bestaan. Leningoverzicht voor details.",
    "kredietbedrag": "BKR is leidend. Leningoverzicht voor actueel saldo.",
    "maandbedrag": "Leningoverzicht is leidend. BKR bevat geen maandbedrag.",
    "saldo": "Leningoverzicht is leidend (actueler dan BKR).",
    "nogAfTeLossen": "Leningoverzicht is leidend.",
    "status": "Keuze adviseur: lopend, aflossen bij/voor passeren.",
    "geldverstrekker": "Hypotheekoverzicht is leidend.",
    "aflosvorm": "Hypotheekoverzicht is leidend.",
    "rentePercentage": "Hypotheekoverzicht is leidend. Check of actueel (na RVP-verlenging).",
    "restschuld": "Hypotheekoverzicht is leidend. Moet recent zijn (<3 mnd).",
}
for row in ws4.iter_rows(min_row=2, max_col=1):
    veld = str(row[0].value or "")
    if veld in regels4:
        c = ws4.cell(row=row[0].row, column=col4, value=regels4[veld])
        c.alignment = Alignment(wrap_text=True, vertical="top")


# --- Voorzieningen ---
ws5 = wb["Voorzieningen & Vermogen"]
col5 = ws5.max_column
regels5 = {
    "type": "Polisblad is leidend voor type verzekering.",
    "aanbieder": "Polisblad is leidend.",
    "premieBedrag": "Polisblad is leidend.",
    "orvDekking": "Polisblad is leidend.",
    "ibanAanvrager": "Bankafschrift is leidend.",
    "ibanPartner": "Bankafschrift is leidend.",
    "spaargeld": "Bankafschrift is leidend. Moet recent zijn (<3 mnd).",
    "eersteJaar": "WGV is leidend voor loondoorbetaling bij ziekte.",
    "tweedeJaar": "WGV is leidend.",
}
for row in ws5.iter_rows(min_row=2, max_col=1):
    veld = str(row[0].value or "")
    if veld in regels5:
        c = ws5.cell(row=row[0].row, column=col5, value=regels5[veld])
        c.alignment = Alignment(wrap_text=True, vertical="top")


wb.save("docs/document-extractie-mapping.xlsx")
print("Validatieregel kolom toegevoegd aan alle 5 tabbladen")
