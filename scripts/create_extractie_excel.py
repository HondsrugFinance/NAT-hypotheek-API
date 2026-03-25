"""Genereer Excel met document-extractie mapping voor fase 2."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2E5644", end_color="2E5644", fill_type="solid")
section_font = Font(bold=True, size=11, color="2E5644")
section_fill = PatternFill(start_color="E3F0E9", end_color="E3F0E9", fill_type="solid")
thin_border = Border(bottom=Side(style="thin", color="E5DFC8"))

def add_header(ws, row, cols):
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

def add_section(ws, row, title, num_cols=4):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = section_font
    cell.fill = section_fill
    for i in range(2, num_cols + 1):
        ws.cell(row=row, column=i).fill = section_fill

def add_row(ws, row, vals):
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="top")

# ============================================================
# TAB 1: Persoonsgegevens
# ============================================================
ws = wb.active
ws.title = "Persoonsgegevens"
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 40
ws.column_dimensions["C"].width = 25
ws.column_dimensions["D"].width = 40

r = 1
add_header(ws, r, ["Veld", "Type / Opties", "Bron document", "Opmerkingen"]); r += 1

add_section(ws, r, "Persoon (aanvrager + partner)"); r += 1
for vals in [
    ("geslacht", "man, vrouw, anders", "Paspoort/ID", ""),
    ("voorletters", "string", "Paspoort/ID", "Afgeleid uit voornamen"),
    ("voornamen", "string", "Paspoort/ID", "Officieel, voluit"),
    ("roepnaam", "string", "", "Niet op document, handmatig"),
    ("tussenvoegsel", "string", "Paspoort/ID", "bijv. van, de, van der"),
    ("achternaam", "string", "Paspoort/ID", ""),
    ("geboortedatum", "datum", "Paspoort/ID", ""),
    ("geboorteplaats", "string", "Paspoort/ID", ""),
    ("geboorteland", "string", "Paspoort/ID", ""),
    ("nationaliteit", "string", "Paspoort/ID", ""),
    ("eerderGehuwd", "boolean", "Echtscheidingsconvenant", ""),
    ("datumEchtscheiding", "datum", "Echtscheidingsconvenant", ""),
    ("weduweWeduwnaar", "boolean", "", "Handmatig"),
]:
    add_row(ws, r, vals); r += 1

r += 1
add_section(ws, r, "Adres & Contact"); r += 1
for vals in [
    ("postcode", "string", "Loonstrook / handmatig", ""),
    ("huisnummer", "string", "Loonstrook / handmatig", ""),
    ("toevoeging", "string", "", ""),
    ("straat", "string", "Loonstrook / handmatig", "Postcode API"),
    ("woonplaats", "string", "Loonstrook / handmatig", "Postcode API"),
    ("land", "string (default: Nederland)", "", ""),
    ("email", "string", "", "Handmatig / e-mail afzender"),
    ("telefoonnummer", "string", "", "Handmatig"),
]:
    add_row(ws, r, vals); r += 1

r += 1
add_section(ws, r, "Identiteit"); r += 1
for vals in [
    ("legitimatiesoort", "paspoort, europese_id, rijbewijs, visum", "Paspoort/ID", ""),
    ("legitimatienummer", "string", "Paspoort/ID", ""),
    ("afgiftedatum", "datum", "Paspoort/ID", ""),
    ("geldigTot", "datum", "Paspoort/ID", ""),
    ("afgifteplaats", "string", "Paspoort/ID", ""),
    ("afgifteland", "string", "Paspoort/ID", ""),
]:
    add_row(ws, r, vals); r += 1

r += 1
add_section(ws, r, "Huishouden"); r += 1
for vals in [
    ("heeftPartner", "boolean", "", "Afgeleid uit aangeleverde docs"),
    ("burgerlijkeStaat", "samenwonend, gehuwd, geregistreerd_partner", "Inschrijving burgerlijke stand", ""),
    ("samenlevingsvorm", "string", "", ""),
    ("kinderen[].geboortedatum", "datum", "", "Handmatig"),
    ("kinderen[].roepnaam", "string", "", "Handmatig"),
]:
    add_row(ws, r, vals); r += 1

# ============================================================
# TAB 2: Inkomen
# ============================================================
ws2 = wb.create_sheet("Inkomen")
ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 40
ws2.column_dimensions["C"].width = 30
ws2.column_dimensions["D"].width = 40

r = 1
add_header(ws2, r, ["Veld", "Type / Opties", "Bron document", "Opmerkingen"]); r += 1

add_section(ws2, r, "Loondienst - Werkgever"); r += 1
for vals in [
    ("naamWerkgever", "string", "WGV / Loonstrook / UWV", "WGV is leidend"),
    ("postcodeWerkgever", "string", "WGV", ""),
    ("adresWerkgever", "string", "WGV", ""),
    ("vestigingsplaats", "string", "WGV", ""),
    ("kvkNummer", "string", "WGV", ""),
]:
    add_row(ws2, r, vals); r += 1

r += 1
add_section(ws2, r, "Loondienst - Dienstverband"); r += 1
for vals in [
    ("soortBerekening", "IBL, WGV, flexibel, arbeidsmarktscan", "", "Keuze adviseur"),
    ("beroepstype", "string", "WGV", ""),
    ("functie", "string", "WGV", ""),
    ("soortDienstverband", "string", "WGV", "Vast / tijdelijk / etc."),
    ("gemiddeldUrenPerWeek", "number", "WGV / Loonstrook", ""),
    ("inDienstSinds", "datum", "WGV", ""),
    ("proeftijd", "string", "WGV", ""),
    ("proeftijdVerstreken", "string", "WGV", ""),
    ("einddatumContract", "datum", "WGV", "Alleen bij tijdelijk"),
    ("loonbeslag", "string", "WGV", ""),
    ("onderhandseLening", "string", "WGV", ""),
    ("directeurAandeelhouder", "string", "WGV", ""),
    ("dienstbetrekkingBijFamilie", "string", "WGV", ""),
]:
    add_row(ws2, r, vals); r += 1

r += 1
add_section(ws2, r, "Loondienst - WGV Inkomen"); r += 1
for vals in [
    ("brutoSalaris", "number", "WGV", ""),
    ("periode", "jaar, maand, week, 4_weken", "WGV", ""),
    ("vakantiegeldPercentage", "number", "WGV", "Default 8%"),
    ("vakantiegeldBedrag", "number", "WGV", ""),
    ("eindejaarsuitkering", "number", "WGV", ""),
    ("onregelmatigheidstoeslag", "number", "WGV", ""),
    ("overwerk", "number", "WGV", ""),
    ("provisie", "number", "WGV", ""),
    ("structureelFlexibelBudget", "number", "WGV", ""),
    ("vebAfgelopen12Maanden", "number", "WGV", "Variabele beloning"),
    ("dertiendeMaand", "number", "WGV", ""),
    ("variabelBrutoJaarinkomen", "number", "WGV", ""),
    ("vastToeslagOpHetInkomen", "number", "WGV", ""),
]:
    add_row(ws2, r, vals); r += 1

r += 1
add_section(ws2, r, "Loondienst - IBL (UWV)"); r += 1
for vals in [
    ("gemiddeldJaarToetsinkomen", "number", "UWV (via IBL-tool)", "Berekend door IBL-tool"),
    ("maandelijksePensioenbijdrage", "number", "Loonstrook", "Input voor IBL-tool"),
    ("aantalWerkgevers", "number", "UWV", ""),
]:
    add_row(ws2, r, vals); r += 1

r += 1
add_section(ws2, r, "Loondienst - Flexibel jaarinkomen"); r += 1
for vals in [
    ("jaar1", "number", "Jaaropgave / IB-aangifte", "Meest recent"),
    ("jaar2", "number", "Jaaropgave / IB-aangifte", ""),
    ("jaar3", "number", "Jaaropgave / IB-aangifte", ""),
]:
    add_row(ws2, r, vals); r += 1

r += 1
add_section(ws2, r, "Onderneming"); r += 1
for vals in [
    ("soort", "inkomen_uit_onderneming, inkomensverklaring", "", ""),
    ("rekenmethode", "regulier, 1_2_3_methode", "", "Keuze adviseur"),
    ("nettoWinstJaar1", "number", "Jaarrekening / IB-aangifte", "Meest recent jaar"),
    ("nettoWinstJaar2", "number", "Jaarrekening / IB-aangifte", ""),
    ("nettoWinstJaar3", "number", "Jaarrekening / IB-aangifte", ""),
    ("bedrijfsnaam", "string", "KVK / Jaarrekening", ""),
    ("kvkNummer", "string", "KVK", ""),
    ("type", "ZZP, ZMP", "KVK", ""),
    ("rechtsvorm", "Eenmanszaak, VOF, BV, etc.", "KVK", ""),
    ("startdatumOnderneming", "datum", "KVK", ""),
]:
    add_row(ws2, r, vals); r += 1

r += 1
add_section(ws2, r, "Pensioen"); r += 1
for vals in [
    ("ouderdomspensioen.bedrag", "number", "Pensioenspecificatie / UPO", ""),
    ("ouderdomspensioen.ingangsdatum", "datum", "Pensioenspecificatie / UPO", ""),
    ("ouderdomspensioen.standPer", "datum", "Pensioenspecificatie / UPO", ""),
    ("partnerpensioen.verzekerdVoor", "number", "Pensioenspecificatie / UPO", ""),
    ("wezenpensioen.verzekerd", "number", "Pensioenspecificatie / UPO", ""),
]:
    add_row(ws2, r, vals); r += 1

r += 1
add_section(ws2, r, "Uitkering"); r += 1
for vals in [
    ("soortUitkering", "ANW, IVA, WGA, Wajong, etc.", "Toekenningsbesluit", ""),
    ("jaarlijksBrutoInkomen", "number", "Betaalspecificatie uitkering", ""),
    ("startdatum", "datum", "Toekenningsbesluit", ""),
    ("einddatum", "datum", "Toekenningsbesluit", ""),
]:
    add_row(ws2, r, vals); r += 1

r += 1
add_section(ws2, r, "Ander inkomen"); r += 1
for vals in [
    ("soortAnderInkomen", "Partneralimentatie, Box 2, Schenking, etc.", "Echtscheidingsconvenant / IB", ""),
    ("jaarlijksBrutoInkomen", "number", "", ""),
    ("startdatum", "datum", "", ""),
    ("einddatum", "datum", "", ""),
]:
    add_row(ws2, r, vals); r += 1

# ============================================================
# TAB 3: Onderpand & Financiering
# ============================================================
ws3 = wb.create_sheet("Onderpand & Financiering")
ws3.column_dimensions["A"].width = 35
ws3.column_dimensions["B"].width = 35
ws3.column_dimensions["C"].width = 30
ws3.column_dimensions["D"].width = 40

r = 1
add_header(ws3, r, ["Veld", "Type / Opties", "Bron document", "Opmerkingen"]); r += 1

add_section(ws3, r, "Onderpand"); r += 1
for vals in [
    ("postcode", "string", "Koopovereenkomst", ""),
    ("huisnummer", "string", "Koopovereenkomst", ""),
    ("straat", "string", "Koopovereenkomst", ""),
    ("woonplaats", "string", "Koopovereenkomst", ""),
    ("typeWoning", "woning, appartement, overig", "Taxatierapport", ""),
    ("soortOnderpand", "2-1-kap, tussenwoning, etc.", "Taxatierapport", ""),
    ("marktwaarde", "number", "Taxatierapport", ""),
    ("marktwaardeNaVerbouwing", "number", "Taxatierapport", ""),
    ("wozWaarde", "number", "WOZ-beschikking", ""),
    ("energielabel", "A++++ t/m G, geen_label", "Energielabel / EP-Online", ""),
    ("erfpacht", "boolean", "Koopovereenkomst", ""),
    ("jaarlijkseErfpacht", "number", "Koopovereenkomst / erfpachtakte", ""),
]:
    add_row(ws3, r, vals); r += 1

r += 1
add_section(ws3, r, "Financieringsopzet (aankoop)"); r += 1
for vals in [
    ("aankoopsomWoning", "number", "Koopovereenkomst", ""),
    ("overdrachtsbelastingPercentage", "1%, 2%, 8%, 10.4%", "", "Afh. van leeftijd/situatie"),
    ("verbouwing", "number", "Verbouwingsspecificatie", ""),
    ("ebv", "number", "", "Energiebesparende voorzieningen"),
    ("taxatiekosten", "number", "Taxatierapport", ""),
    ("eigenGeld", "number", "Bankafschrift", ""),
]:
    add_row(ws3, r, vals); r += 1

r += 1
add_section(ws3, r, "Koopovereenkomst specifiek"); r += 1
for vals in [
    ("koopprijs", "number", "Koopovereenkomst", "= aankoopsomWoning"),
    ("leveringsdatum", "datum", "Koopovereenkomst", ""),
    ("datumBankgarantie", "datum", "Koopovereenkomst", ""),
    ("ontbindendeVoorwaardenDatum", "datum", "Koopovereenkomst", ""),
]:
    add_row(ws3, r, vals); r += 1

# ============================================================
# TAB 4: Verplichtingen & Hypotheken
# ============================================================
ws4 = wb.create_sheet("Verplichtingen & Hypotheken")
ws4.column_dimensions["A"].width = 30
ws4.column_dimensions["B"].width = 35
ws4.column_dimensions["C"].width = 30
ws4.column_dimensions["D"].width = 40

r = 1
add_header(ws4, r, ["Veld", "Type / Opties", "Bron document", "Opmerkingen"]); r += 1

add_section(ws4, r, "Verplichtingen (per verplichting)"); r += 1
for vals in [
    ("type", "doorlopend, aflopend, lease, studie, alimentatie", "BKR / Leningoverzicht", ""),
    ("kredietnummer", "string", "BKR / Leningoverzicht", ""),
    ("maatschappij", "string", "BKR / Leningoverzicht", ""),
    ("kredietbedrag", "number", "BKR / Leningoverzicht", ""),
    ("maandbedrag", "number", "Leningoverzicht", ""),
    ("saldo", "number", "Leningoverzicht", ""),
    ("nogAfTeLossen", "number", "Leningoverzicht", ""),
    ("rentepercentage", "number", "Leningoverzicht", ""),
    ("ingangsdatum", "datum", "BKR / Leningoverzicht", ""),
    ("einddatum", "datum", "BKR / Leningoverzicht", ""),
    ("status", "lopend, aflossen bij/voor passeren", "", "Keuze adviseur"),
]:
    add_row(ws4, r, vals); r += 1

r += 1
add_section(ws4, r, "Bestaande hypotheek - inschrijving"); r += 1
for vals in [
    ("geldverstrekker", "string", "Hypotheekoverzicht", ""),
    ("inschrijving", "number", "Hypotheekoverzicht", ""),
    ("rangorde", "number", "Hypotheekoverzicht", ""),
    ("nhg", "boolean", "Hypotheekoverzicht", ""),
]:
    add_row(ws4, r, vals); r += 1

r += 1
add_section(ws4, r, "Bestaande hypotheek - leningdeel"); r += 1
for vals in [
    ("aflosvorm", "annuitair, lineair, aflossingsvrij, etc.", "Hypotheekoverzicht", ""),
    ("bedrag", "number", "Hypotheekoverzicht", "Oorspronkelijke hoofdsom"),
    ("rentePercentage", "number", "Hypotheekoverzicht", ""),
    ("ingangsdatum", "datum", "Hypotheekoverzicht", ""),
    ("looptijd", "number (maanden)", "Hypotheekoverzicht", ""),
    ("einddatum", "datum", "Hypotheekoverzicht", ""),
    ("renteVastPeriode", "number (jaren)", "Hypotheekoverzicht", ""),
    ("einddatumRvp", "datum", "Hypotheekoverzicht", ""),
    ("restschuld", "number", "Hypotheekoverzicht", ""),
]:
    add_row(ws4, r, vals); r += 1

# ============================================================
# TAB 5: Voorzieningen & Vermogen
# ============================================================
ws5 = wb.create_sheet("Voorzieningen & Vermogen")
ws5.column_dimensions["A"].width = 30
ws5.column_dimensions["B"].width = 35
ws5.column_dimensions["C"].width = 30
ws5.column_dimensions["D"].width = 40

r = 1
add_header(ws5, r, ["Veld", "Type / Opties", "Bron document", "Opmerkingen"]); r += 1

add_section(ws5, r, "Verzekeringen (per verzekering)"); r += 1
for vals in [
    ("type", "ORV, AOV, woonlasten, lijfrente, leven", "", "Keuze adviseur / polis"),
    ("eigenaar", "aanvrager, partner, gezamenlijk", "", ""),
    ("aanbieder", "string", "Polisblad", ""),
    ("polisnummer", "string", "Polisblad", ""),
    ("ingangsdatum", "datum", "Polisblad", ""),
    ("einddatum", "datum", "Polisblad", ""),
    ("soortDekking", "gelijkblijvend, annuitair, lineair", "Polisblad", ""),
    ("premieBedrag", "number", "Polisblad", ""),
    ("orvDekking", "number", "Polisblad", "Per verzekerde"),
    ("dekkingAO", "number", "Polisblad", ""),
    ("dekkingWW", "number", "Polisblad", ""),
]:
    add_row(ws5, r, vals); r += 1

r += 1
add_section(ws5, r, "Werkgeversverzekeringen"); r += 1
for vals in [
    ("type", "WGA-hiaat, WIA-excedent, loondoorbetaling", "WGV / loonstrook", ""),
    ("eersteJaar", "percentage", "WGV", "Loondoorbetaling bij ziekte"),
    ("tweedeJaar", "percentage", "WGV", ""),
]:
    add_row(ws5, r, vals); r += 1

r += 1
add_section(ws5, r, "Vermogen"); r += 1
for vals in [
    ("ibanAanvrager", "string", "Bankafschrift", ""),
    ("ibanPartner", "string", "Bankafschrift", ""),
    ("spaargeld", "number", "Bankafschrift / Vermogensoverzicht", ""),
    ("beleggingen", "number", "Vermogensoverzicht", ""),
    ("schenking", "number", "", "Handmatig"),
    ("eigenaar", "gezamenlijk, aanvrager, partner", "", ""),
]:
    add_row(ws5, r, vals); r += 1

# ============================================================
# TAB 6: Document naar Veld mapping
# ============================================================
ws6 = wb.create_sheet("Document naar Veld")
ws6.column_dimensions["A"].width = 30
ws6.column_dimensions["B"].width = 55
ws6.column_dimensions["C"].width = 15
ws6.column_dimensions["D"].width = 45

r = 1
add_header(ws6, r, ["Document", "Vult deze velden", "Prioriteit", "Opmerkingen"]); r += 1

docs = [
    ("Paspoort / ID-kaart",
     "geslacht, voorletters, voornamen, tussenvoegsel, achternaam, geboortedatum, geboorteplaats, nationaliteit, legitimatiesoort, legitimatienummer, afgiftedatum, geldigTot, afgifteplaats",
     "1 (hoogste)", "Officieel document, leidend voor namen"),
    ("Werkgeversverklaring",
     "naamWerkgever, adres, functie, soortDienstverband, inDienstSinds, brutoSalaris, vakantiegeld, eindejaarsuitkering, ORT, overwerk, provisie, 13e maand, proeftijd, loonbeslag, onderhandseLening",
     "1 (hoogste)", "Leidend voor inkomen loondienst"),
    ("Loonstrook",
     "brutoMaandloon, werkgever, periode, bijzondere beloningen, adres werknemer, pensioenbijdrage",
     "2 (controle)", "Verificatie van WGV, bron voor pensioenbijdrage IBL"),
    ("UWV Verzekeringsbericht",
     "gemiddeldJaarToetsinkomen (via IBL-tool), aantalWerkgevers",
     "1 (hoogste)", "IBL-tool berekent toetsinkomen"),
    ("Koopovereenkomst",
     "adres onderpand, aankoopsomWoning, leveringsdatum, bankgarantie datum, ontbindende voorwaarden, erfpacht",
     "1 (hoogste)", "Leidend voor aankoopgegevens"),
    ("Taxatierapport",
     "marktwaarde, marktwaardeNaVerbouwing, typeWoning, soortOnderpand, taxatiedatum",
     "1 (hoogste)", "Leidend voor woningwaarde"),
    ("Energielabel",
     "energielabel, afgiftedatumEnergielabel",
     "1", "Ook via EP-Online API"),
    ("Hypotheekoverzicht",
     "geldverstrekker, bedrag, restschuld, rentePercentage, maandlast, einddatumRvp, aflosvorm",
     "1", "Bestaande hypotheek"),
    ("Bankafschrift",
     "IBAN, saldo, bank",
     "1", "Vermogen + incassorekening"),
    ("Vermogensoverzicht",
     "totaalVermogen, beleggingen",
     "1", ""),
    ("BKR",
     "type verplichting, kredietnummer, maatschappij, kredietbedrag, saldo",
     "1", ""),
    ("Leningoverzicht",
     "maandbedrag, restschuld, rentepercentage, einddatum",
     "1", "Aanvulling op BKR"),
    ("Jaarrekening",
     "nettoWinstJaar1/2/3, bedrijfsnaam, rechtsvorm",
     "1", "Ondernemers"),
    ("IB-aangifte",
     "nettoWinstJaar1/2/3, box 2 inkomen",
     "2", "Controle op jaarrekening"),
    ("IB60",
     "vastgesteld inkomen",
     "1", "Definitief, van Belastingdienst"),
    ("Echtscheidingsconvenant",
     "datumEchtscheiding, partneralimentatie, kinderalimentatie, verdeling vermogen",
     "1", ""),
    ("Pensioenspecificatie / UPO",
     "ouderdomspensioen bedrag + datum, partnerpensioen, wezenpensioen",
     "1", ""),
    ("Toekenningsbesluit uitkering",
     "soortUitkering, startdatum, einddatum, bedrag",
     "1", ""),
    ("Verbouwingsspecificatie",
     "verbouwingskosten, specificatie per post",
     "1", ""),
]

for doc, velden, prio, opm in docs:
    add_row(ws6, r, [doc, velden, prio, opm]); r += 1

out = "docs/document-extractie-mapping.xlsx"
wb.save(out)
print(f"Excel aangemaakt: {out}")
