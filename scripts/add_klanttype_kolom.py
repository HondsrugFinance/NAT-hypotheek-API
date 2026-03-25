"""Voeg kolom 'Vereist bij klanttype' toe aan tabblad 'Document naar Veld'."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.load_workbook("docs/document-extractie-mapping.xlsx")

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2E5644", end_color="2E5644", fill_type="solid")

ws = wb["Document naar Veld"]

# Voeg kolom toe na "Prioriteit" (kolom C), schuif "Opmerkingen" op
# Huidige kolommen: Document | Vult deze velden | Prioriteit | Opmerkingen
# Gewenst: Document | Vult deze velden | Prioriteit | Vereist bij klanttype | Opmerkingen

# Insert kolom D (schuift bestaande D naar E)
ws.insert_cols(4)

# Header
cell = ws.cell(row=1, column=4, value="Vereist bij klanttype")
cell.font = header_font
cell.fill = header_fill
cell.alignment = Alignment(horizontal="left")
ws.column_dimensions["D"].width = 45

# Klanttypen: AB=Aankoop Bestaand, ANP=Aankoop Nieuwbouw Project,
# ANEB=Aankoop Nieuwbouw Eigen Beheer, VH=Verhogen, OS=Oversluiten, UP=Uitkoop Partner
# "Alle" = bij elk klanttype

klanttype_mapping = {
    "Paspoort / ID-kaart": "Alle",
    "Werkgeversverklaring": "Alle (bij loondienst)",
    "Loonstrook": "Alle (bij loondienst)",
    "UWV Verzekeringsbericht": "Alle (bij loondienst)",
    "Koopovereenkomst": "Aankoop (BB, NP, NEB)",
    "Taxatierapport": "Alle",
    "Energielabel": "Alle",
    "Hypotheekoverzicht": "Doorstromer, Verhoger, Oversluiter, Uitkoop",
    "Bankafschrift": "Alle",
    "Vermogensoverzicht": "Alle",
    "BKR": "Alle",
    "Leningoverzicht": "Alle (bij lopende leningen)",
    "Jaarrekening": "Alle (bij ondernemer/DGA)",
    "IB-aangifte": "Alle (bij ondernemer/DGA)",
    "IB60": "Alle (bij ondernemer/DGA)",
    "Echtscheidingsconvenant": "Uitkoop partner",
    "Pensioenspecificatie / UPO": "Alle (bij pensioen/AOW)",
    "Toekenningsbesluit uitkering": "Alle (bij uitkering)",
    "Verbouwingsspecificatie": "Verhoger, Aankoop (bij verbouwing)",
}

for row in ws.iter_rows(min_row=2, max_col=1):
    doc = str(row[0].value or "")
    if doc in klanttype_mapping:
        c = ws.cell(row=row[0].row, column=4, value=klanttype_mapping[doc])
        c.alignment = Alignment(wrap_text=True, vertical="top")

wb.save("docs/document-extractie-mapping.xlsx")
print("Kolom 'Vereist bij klanttype' toegevoegd aan Document naar Veld tabblad")
